#!/usr/bin/env python
# coding: utf-8

# This project contains the details scrapped from Rotten Tomato to extract the required data and store it as an excel sheet. Webscrapping allows us to collect data from websites that do not provide APIs or structured data formats. Web scraping provides access to data that can be analyzed to gain insights, identify patterns, and make data-driven decisions. 

# # WEB SCRAPPING ROTTEN TOMATOES WEBSITE USING PYTHON

# Here we focus on scrapping the web details on `Rotten Tomato` website which contains top 105 films that has 100% score.Details such as Title of the Movie, Year of Release and the Director of the movie are scrapped from the website to obtain the required datasheet. `BeautifulSoup` is used for importing and inspecting the entire `HTML` code of the website to obtain the necessary details by performing certain operations using `Python`

# In[1]:


#importing the necessary libraries
from bs4 import BeautifulSoup
import openpyxl
import requests


# After importing the libraries, the link of the `Rotten Tomato` website is obtained and assigned to the variable named as `response` using the requests library. The `raise_for_status()` function is mentioned in order to obtain the correct URL form. If the URL being wrong it throws as error as `404` client ereor

# In[2]:


response = requests.get('https://editorial.rottentomatoes.com/guide/movies-100-percent-score-rotten-tomatoes')
response.raise_for_status()


# In[3]:


print(response)


# The response code <Response [200]> indicates that the request to the URL was successful. 
# In the context of web scraping, a response code of 200 means that the web page was retrieved successfully, and you can proceed to extract the desired data from the response content.
# Other responses such as 404 for "Not Found," 403 for "Forbidden," or 500 for "Internal Server Error." might occur if the web page is not being retrieved

# In[4]:


#creating an excel workbook and creating the required columns
excel=openpyxl.Workbook()
print(excel.sheetnames)
sheet=excel.active
sheet.title='Rotten Tomato Top Rated Movies'
print(excel.sheetnames)
sheet.append(['Movie Name','Released Year','Director'])


# The details being scrapped must be saved in an excel sheet. Inorder to achieve that `openpyxl` library is used for creating a exce workbook. 
# 
# The sheet name of the excel is assigned as the features being extracted are given the column title of the excel.

# In[5]:


#importing BeautifulSoup library for obtaining HTML of the website
soup=BeautifulSoup(response.text,'html.parser')
print(soup)


# By using BeautifulSoup, the HTML code is stored in the `soup` variable which is used to obtain certain details from the code for generating the required excel

# In[6]:


movies=soup.find('div', class_="articleContentBody").find_all('div',class_='row countdown-item')
print(len(movies))


# By using the find function, the movies under the class `articleContentBody` are displayed by providing an additional class `row countdown-item` which specially contains the movie name in the HTML code.

# In[7]:


for mv in movies:
    name=mv.find('div',class_="article_movie_title").a.text
    year=mv.find('div',class_="article_movie_title").span.text.strip('()')
    dr=mv.find('div',class_="info director").a.text
    print(name,year,dr)
    sheet.append([name,year,dr])


# By using the `For-Loop` condition, the Title, Year of release and Director details are obtained for the entire number of 105 movies by mentioning the corresponding HTML Tag.
# 
# - "article_movie_title" -  for obtaining the title of the movie
# - "article_movie_title" -  for obataining the Year of release(scrapping of the paranthesis)
# - "info director"       -  to obtain the name of the director
# 
# As the title of the movie and the year of release are locted inside the same tag, the same ta name has been used

# In[8]:


excel.save('Rotten Tomato Ratings.xlsx')


# After scrapping off the necessary details, they are stored in the created excel workbook named as `Rotten Tomato Top Rated Movies` with their corresponding feature names an is saved.
